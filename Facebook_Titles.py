
'''
 _______  _______  _______  _______  _______  _______  _______  ___   _    _______  ___   _______  ___      _______  _______ 
|       ||   _   ||       ||       ||  _    ||       ||       ||   | | |  |       ||   | |       ||   |    |       ||       |
|    ___||  |_|  ||       ||    ___|| |_|   ||   _   ||   _   ||   |_| |  |_     _||   | |_     _||   |    |    ___||  _____|
|   |___ |       ||       ||   |___ |       ||  | |  ||  | |  ||      _|    |   |  |   |   |   |  |   |    |   |___ | |_____ 
|    ___||       ||      _||    ___||  _   | |  |_|  ||  |_|  ||     |_     |   |  |   |   |   |  |   |___ |    ___||_____  |
|   |    |   _   ||     |_ |   |___ | |_|   ||       ||       ||    _  |    |   |  |   |   |   |  |       ||   |___  _____| |
|___|    |__| |__||_______||_______||_______||_______||_______||___| |_|    |___|  |___|   |___|  |_______||_______||_______|

'''


'''
Hey all!

So this here webscraping script isn't actually for webscraping! It builds off another guys webscraping tool at this Github page:

https://github.com/minimaxir/facebook-page-post-scraper

If you set up the facebook app ID from his site, you'll have access to the analysis of any facebook page!
So this script takes the files that get output of his function and processes them for emotional content as well as objective metrics like verb/noun usage.

All you need to do is replace your file name right here ( in my_file_name ):
'''
global my_file_name
my_file_name = 'NYT Articles.xlsx'
'''
Then, make sure you have the right modules, and kick back and relax as the script does all the hard analysis (even correlations too)!
'''

# Standard Libaries
import jdcal
import datetime
import urllib2
import collections
import urllib
import re,os,csv,time
from nltk.tag import pos_tag
import math
import numpy

# Non-Standard Libaries
import requests
from BeautifulSoup import BeautifulSoup
import bs4
import google
from google import search
import xlrd
from openpyxl import Workbook
from textblob import TextBlob
from matplotlib.mlab import PCA as mlabPCA

def dotproduct(v1, v2):
  return sum((a*b) for a, b in zip(v1, v2))

def length(v):
  return math.sqrt(dotproduct(v, v))

def angle(v1, v2):
    try:
        return str(math.acos(dotproduct(v1, v2) / (length(v1) * length(v2))))
    except:
        return 'NaN'

def Run():
    background = Title_Background(my_file_name)
    Popular_Titles(background,my_file_name[:-5]+'_Analysis.xlsx')
      
            
def GeneralExcel(data,fname):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = datetime.datetime.now()
    for d in data:
        ws.append(d)
    wb.save(fname)

def Dictionaries():
    global tags, AFINN_key, AFINN_value
    tags = [('CC','coordinating conjunction'),
    ('CD','cardinal number'),
    ('DT','determiner'),
    ('EX','	existential there'),
    ('FW','	foreign word'),
    ('IN','	preposition/subordinating conjunction'),
    ('JJ','	adjective'),
    ('JJR','adjective, comparative'),
    ('JJS','adjective, superlative'),
    ('LS','	list marker'),
    ('MD','	modal'),
    ('NN','	noun, singular or mass'),
    ('NNS','noun plural'),
    ('NNP','proper noun, singular'),
    ('NNPS','proper noun, plural'),
    ('PDT','predeterminer'),
    ('POS','possessive ending'),
    ('PRP','personal pronoun'),
    ('PRP$','possessive pronoun'),
    ('RB','adverb'),
    ('RBR','adverb, comparative'),
    ('RBS','adverb, superlative'),
    ('RP ','particle'),
    ('TO ','to'),
    ('UH ','interjection'),
    ('VB ','verb, base form'),
    ('VBD','verb, past tense'),
    ('VBG','verb, gerund/present participle'),
    ('VBN','verb, past participle'),
    ('VBP','verb, sing. present, non-3d'),
    ('VBZ','verb, 3rd person sing. present'),
    ('WDT','wh-determiner'),
    ('WP ','wh-pronoun'),
    ('WP$','possessive wh-pronoun'),
    ('WRB','wh-abverb')]
    with open('AFINN.csv', 'rb') as f:
        reader = csv.reader(f)
        AFINN = list(reader)
    AFINN_key = [i[0].replace('\t','') for i in AFINN]
    AFINN_value = [int(i[1]) for i in AFINN]

def AFINN_Analysis(text):
    text = ' '.join(text)
    text = text.lower().split(' ')
    key = [-5,-4,-3,-2,-1,0,1,2,3,4,5]
    value = [0,0,0,0,0,0,0,0,0,0,0]
    text = [i.encode('ascii','ignore') for i in text]
    for word in text:
        if word in AFINN_key:
            value[key.index(AFINN_value[AFINN_key.index(word)])] += 1
    return value,len(text)

# Purpose: This function opens title databases with specific formatting for results
def Title_Background(path):
    background,workbook = [],xlrd.open_workbook(path)
    for sheet in workbook.sheets():
        for row in range(1,sheet.nrows):
            background.append([(sheet.cell(row,2).value)])
            for col in [6,7,8]:
                background[-1].append(float(sheet.cell(row,col).value))
    background.sort(key=lambda x: -x[1])  # Sort by number of likes
    return background

def Word_ID(background):
    WORDS,IDS,text = [],[],[]
    for title in [i[0] for i in background]:
        try:
            tagged_title = pos_tag(title.split())
            words,ids = [i[0] for i in tagged_title],[i[1] for i in tagged_title]
            IDS += ids
            text.append(title)
        except:
            print 'Title rejected due to unknown data format.'
    counter=collections.Counter(IDS)
    return counter.keys(),counter.values(),text

def Popular_Titles(background,fname):
    iteration = 250
    wb = Workbook()
    ws = wb.active

    spectrum = ['-5','-4','-3','-2','-1','0','1','2','3','4','5']
    angles = ['','','']
    
    ws['A1'] = datetime.datetime.now()
    ws.append(['','','Correlation:']+[i[0] for i in tags])
    ws.append(['Rank Range','Like Range','Average Likes']+[i[1] for i in tags]+['Net Emotion','Emotional Magnitude','Subjectivity','Polarity','Words/Title','','All words'])
    results,pca_data = [],[]
    for stack in range(0,len(background),iteration):
        print 'Starting analysis on titles:',str(stack+1)+'-'+str(stack+iteration)
        if stack + iteration - 1 < len(background) - 1:
            results.append([str(stack+1)+'-'+str(stack+iteration),str(int(background[stack+iteration-1][1]))+'-'+str(int(background[stack][1]))])
            results[-1].append(sum([i[1] for i in background[stack:stack+iteration-1]])/(iteration-1))
        else:
            results.append([str(stack+1)+'-'+str(len(background)),str(int(background[-1][1]))+'-'+str(int(background[stack][1]))])
            results[-1].append(sum([i[1] for i in background[stack:]])/(len(background)-stack+1))
        bg = background[stack:stack+iteration]
        keys,values,text = Word_ID(bg)
        emo,count = AFINN_Analysis(text)
        for tag,temp in tags:
            if tag in keys:
                results[-1].append(float(values[keys.index(tag)])/count)
            else:
                results[-1].append(0)
        words = TextBlob('. '.join(text))
        results[-1] += [float(sum([int(i)*j for i,j in zip(spectrum,emo)]))/iteration,float(sum([abs(int(i))*j for i,j in zip(spectrum,emo)]))/iteration,words.sentiment.subjectivity,words.sentiment.polarity,float(count)/iteration]
        results[-1] += ['','. '.join(text)]

    for i in range(3,len(results[-1])-2):
        t = [results[j][i] for j in xrange(len(results))]
        l = [results[j][2] for j in xrange(len(results))]
        test = [j for j,jj in zip(t,l) if jj > 0]
        like = [math.log10(float(jj)) for j,jj in zip(t,l) if jj > 0]
        temp = numpy.corrcoef(test,like)
        angles.append(str(temp[0][1]))

    ws.append(angles)
    for i in results:
        ws.append(i)
    wb.save(fname)
    print 'Analysis completed!'
    
def polyfit(x, y, degree = 1):
    results = {}

    coeffs = numpy.polyfit(x, y, degree)

     # Polynomial Coefficients
    results['polynomial'] = coeffs.tolist()

    # r-squared
    p = numpy.poly1d(coeffs)
    # fit values, and mean
    yhat = p(x)                         # or [p(z) for z in x]
    ybar = numpy.sum(y)/len(y)          # or sum(y)/len(y)
    ssreg = numpy.sum((yhat-ybar)**2)   # or sum([ (yihat - ybar)**2 for yihat in yhat])
    sstot = numpy.sum((y - ybar)**2)    # or sum([ (yi - ybar)**2 for yi in y])
    results['determination'] = ssreg / sstot

    return str(results)
    

def SearchResults(query,total):
    results = []
    for url in google.search(query, num=25, stop=1):
        results.append(url)
    return results

class WebScrape:
    def __init__(self,query,article_total = 100):
        self.article_total = article_total
        self.nbhd = 5
        check = False
        
        self.urls = SearchResults(query,self.article_total)
        self.words,self.hs_words,self.titles,self.raw_words = [],[],[],[]
        
        hotspots = [['gmo'],['genetically','modified','organism'],['genetically','modified','food']]
        for i,url in enumerate(self.urls):
            print 'Loading site: %d \n URL: %s' % (i,url)
            filename,extension = os.path.splitext('/path/to/somefile.ext')
            if extension == '.pdf':
                print 'PDF file rejected due to formatting problems.'
            else:
                self.URL_Format(url,i)
        if check:
            keepers = []
            for title in self.titles:
                try:
                    print 'Title:',title[0].encode('ascii','ignore')
                    temp = raw_input('Rename: ')
                    if len(temp) > 0:
                        keepers.append([temp,title[1]])
                        print 'Saved!'
                        time.sleep(0.2)
                    else:
                        print 'Rejected.'
                        time.sleep(0.2)
                except:
                    print 'That was weird. Skipping entry.'
            Popular_Titles(keepers,query+'.xlsx')
        else:
            Popular_Titles(self.titles,query+'.xlsx')
        self.Hotspot_Analysis(self.words[-1],hotspots)
        #Publish_TXT('Hotspot words.txt',' '.join(self.hs_words))

    def URL_Format(self,url,index):  
        try:
            opener = urllib2.build_opener()
            print 'Opener.'
            opener.addheaders = [('User-agent','Google Chrome')]
            print 'Add headers.'
            response = opener.open(url)
            print 'Open URL.'
            page = response.read()
            print 'Response.'
            soup = BeautifulSoup(page)
            print 'Formatting webppage titled: %s' % soup.title.string
            texts = soup.findAll(text=True)
            visible_texts = filter(visible, texts)
            visible_texts = ' '.join(visible_texts)
            visible_texts = visible_texts.replace('\n',' ').replace('\t',' ')
            self.raw_words.append(visible_texts)
            
            temp = ' '.join(visible_texts.split())
            temp = Remove_Characters(temp,""".?(),'":;^[]""")
            temp = [word.lower() for word in temp.split(' ')]
            words = [x for x in temp if not any(c.isdigit() for c in x)]
            self.titles.append([soup.title.string,index])
            self.words.append(words)
        except:
            print 'This website is acting very strange, entry rejected.'
        
    def Hotspot_Analysis(self,words,hotspots):
        for hotspot in hotspots:
            lp = len(hotspot)
            self.hs_words += list(words[j+i] for j in (range(-self.nbhd,0) + range(lp,self.nbhd+lp)) for i,x in enumerate(words[:-(lp+self.nbhd)]) if words[i:i+lp] == hotspot)
            #print 'Singular search:',len(self.hs_words)
            self.hs_words += list(words[j+i] for j in (range(-self.nbhd,0) + range(lp,self.nbhd+lp)) for i,x in enumerate(words[:-(lp+self.nbhd)]) if words[i:i+lp] == hotspot[:-1]+[hotspot[-1]+'s'])
            #print 'Plural search:',len(self.hs_words)

def visible(element):
    if element.parent.name in ['style', 'script', '[document]', 'head', 'title']:
        return False
    elif re.match('<!--.*-->', str(element)):
        return False
    return True

def Remove_Characters(string,chars):
    for char in chars:
        string = string.replace(char,'')
    return string

def Publish_TXT(fname,data):
    with open(fname, 'w') as text_file:
        text_file.write(data.encode('utf8'))

Dictionaries()
Run()
